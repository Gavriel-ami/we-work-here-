import openpyxl

# def delete_rows_from_line(input_file, sheet_name, number_of_lines):
#     # Load the workbook and sheet
#     workbook = openpyxl.load_workbook(input_file)
#     sheet = workbook[sheet_name]

#     # Calculate the starting row to delete
#     start_delete_row = number_of_lines + 1

#     # Delete rows from start_delete_row to the end
#     while sheet.max_row >= start_delete_row:
#         sheet.delete_rows(start_delete_row, amount=1)

#     # Save the modified workbook to the same input file
#     workbook.save(input_file)
#     print(f"Rows deleted starting from line {start_delete_row}. Changes saved to {input_file}.")

# # Example usage
# input_file = 'Batch_Control_History_1.xlsx'
# sheet_name = 'Sheet1'
# number_of_lines = 241

# delete_rows_from_line(input_file, sheet_name, number_of_lines)

def delete_rows_from_line(input_file, sheet_name, number_of_lines):
    # Load the workbook and sheet
    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook[sheet_name]

    # Calculate the starting row to delete
    start_delete_row = number_of_lines + 1

    # Use a more efficient way to delete rows in batch
    max_row = sheet.max_row
    sheet.delete_rows(start_delete_row, max_row - start_delete_row + 1)

    # Save the modified workbook to the same input file
    workbook.save(input_file)
    print(f"Rows deleted starting from line {start_delete_row}. Changes saved to {input_file}.")
