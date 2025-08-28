import ast
import os 
import sys 
import time
import shutil
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font , numbers
from datetime import datetime, timedelta

from big_tindex_values import generate_big_tindex_values
from clean_xl import delete_rows_from_line
from copy_file import copy_file_to_same_folder
from func_tindex_duplicate_point import update_excel_column
from func_change_dt import update_excel_column
from rename_and_copy import rename_file_and_copy

  
  
def timer_func(func): 
    # This function shows the execution time of  
    # the function object passed 
    def wrap_func(*args, **kwargs): 
        t1 = time() 
        result = func(*args, **kwargs) 
        t2 = time() 
        print(f'Function {func.__name__!r} executed in {(t2-t1):.4f}s') 
        return result 
    return wrap_func 

# def copy_template_to_destination():
#     # Define the source file path, destination directory, and new file name
#     src_file = r'C:\projects\AMI WE WORK HERE\AmiMult\python_scripts\B_C_History-main\_Batch_Control_History_template.xlsx'
#     dest_directory = r'C:\projects\AMI WE WORK HERE'
#     new_file_name = "Batch_Control_History.xlsx"
#     return src_file,dest_directory,new_file_name

def copy_template_to_destination():
    current_dir = os.getcwd()
    # Define the source file path relative to the current directory (three levels up)
    src_file = os.path.join(current_dir, os.pardir, os.pardir, os.pardir, "AmiMult", "python_scripts", "B_C_History-main", "_Batch_Control_History_template.xlsx")
    dest_directory = os.path.abspath(os.path.join(current_dir, os.pardir, os.pardir, os.pardir))
    new_file_name = "Batch_Control_History.xlsx"
    new_file_path = os.path.join(dest_directory, new_file_name)
    shutil.copyfile(src_file, new_file_path)
    return src_file, dest_directory, new_file_name

def run_flow_build_batch_hitory(batch_file_template,jump_time,hours_back_to_test,dt_values):
    batch_control_history_black = 'Batch_Control_History_1.xlsx'
    number_of_lines = build_history_batch(batch_file_template,batch_control_history_black,jump_time,hours_back_to_test,dt_values) 
    return number_of_lines

def generate_dynamic_data(unique_times_count, repetition_count):
    from datetime import datetime, timedelta

    # Generate base timestamps dynamically
    base_time = datetime.strptime('6/6/2024 9:00', '%m/%d/%Y %H:%M')
    base_data = [(base_time + timedelta(hours=i)).strftime('%m/%d/%Y %H:%M') for i in range(unique_times_count)]

    data = {
        'Max Date': [time for time in base_data for _ in range(repetition_count)]
    }
    return data

def build_history_batch(input_file,batch_control_history,jump_time,hours_back_to_test,dt_values):
    # Timer start
    start_time = time.time()
    # Load the existing Excel file
    print(start_time)
    df_existing = pd.read_excel(input_file)
    # Build the `data` for the seting the times. 
    unique_times_count = hours_back_to_test * jump_time  # Number of unique timestamps you want
    repetition_count = 1  # Number of times each timestamp should be repeated
    data = generate_dynamic_data(unique_times_count, repetition_count)
    df = pd.DataFrame(data)
    # Get the current time and set seconds to zero
    current_time = datetime.now().replace(second=0, microsecond=0)
    print(current_time.strftime('%m/%d/%Y  %I:%M:%S %p'))  # Print the current time in the desired format

    
#    Iterate through the DataFrame and adjust times
    for i in range(len(df)):
        hours_to_subtract = (i // jump_time)  # Adjust to jump over x rows that is set in jump_time var.
        new_time = current_time - timedelta(hours=hours_to_subtract)
        df.at[i, 'Max Date'] = new_time.strftime('%m/%d/%Y  %I:%M:%S %p').upper()  # Add AM/PM in capital letters
        print(df['Max Date'])

    # Concatenate the DataFrame with itself 8 times
    df_concatenated = pd.concat([df] * len(dt_values), ignore_index=True)

    # Ensure the existing DataFrame has enough rows to be updated
    if len(df_existing) < len(df_concatenated):
        raise ValueError("The existing DataFrame does not have enough rows to accommodate the new data.")

    # Update the 'Max Date' column of the existing DataFrame
    df_existing['Max Date'] = df_concatenated['Max Date']
    df_existing.to_excel(batch_control_history, index=False)
    print(f"completed the building of the {batch_control_history}")

    #  Timer end
    end_time = time.time()
    elapsed_time = end_time - start_time
    # Print the elapsed time
    print(f"Elapsed time: {elapsed_time:.2f} seconds")

    return (df.shape[0] * len(dt_values))

def process_files(batch_control_history, batch_file_template, sheet_name='Control Sheet', number_of_lines=0):
    start_time = time.time()
    
    # Read the data from the files
    df = pd.read_excel(batch_control_history, sheet_name='Sheet1')
    df1 = pd.read_excel(batch_file_template, sheet_name=sheet_name)
    
    # Convert 'Max Date' in df to datetime
    df['Max Date'] = pd.to_datetime(df['Max Date'], errors='coerce')
    
    # Columns to be updated
    columns_to_update = ['Max Date', 'Tindex Duplicate  Point', 'DT ']
    
    # Load the workbook and sheet
    workbook = load_workbook(batch_file_template)
    sheet = workbook[sheet_name]
    
    # Define the font
    font = Font(name='Calibri', size=11)
    
    # Update columns
    for col in columns_to_update:
        col_index = df1.columns.get_loc(col) + 1  # openpyxl is 1-indexed
        for row_index, (value1, value2) in enumerate(zip(df1[col], df[col]), start=2):
            cell = sheet.cell(row=row_index, column=col_index)
            cell.value = value2
            cell.font = font

        # Ensure the column is visible and set a standard width
        column_letter = sheet.cell(row=1, column=col_index).column_letter
        sheet.column_dimensions[column_letter].hidden = False
        sheet.column_dimensions[column_letter].width = 20  # Adjust width for the new format

    update_columns_time = time.time()
    print(f"read_data_and_update_columns executed in {update_columns_time - start_time} seconds.")
    
    # Delete rows
    if number_of_lines > 0:
        start_delete_row = number_of_lines + 1
        max_row = sheet.max_row
        sheet.delete_rows(start_delete_row, max_row - start_delete_row + 1)
        print(f"Rows deleted starting from line {start_delete_row}.")
    
    delete_rows_time = time.time()
    print(f"delete_rows_from_line executed in {delete_rows_time - update_columns_time} seconds.")

    # Save the workbook
    workbook.save(batch_file_template)
    print(f"Workbook '{batch_file_template}' saved successfully.")
    
    end_time = time.time()
    print(f"Total execution time: {end_time - start_time} seconds.")

def casting_arguments(tindex_values_str, hours_back_to_test_str, dt_values_str):
    """
    Process command line arguments and convert them to appropriate data types.
    
    Parameters:
        tindex_values_str (str): String representation of a list of tindex values.
        hours_back_to_test_str (str): String representation of hours back to test.
        dt_values_str (str): String representation of a list of dt values.
        
    Returns:
        tuple: (list of floats, int, list of floats)
    """
    try:
        # Convert string representations to actual types
        tindex_values = ast.literal_eval(tindex_values_str)
        hours_back_to_test = int(hours_back_to_test_str)
        dt_values = ast.literal_eval(dt_values_str)

        # Return the processed values
        return tindex_values, hours_back_to_test, dt_values

    except (ValueError, SyntaxError) as e:
        print(f"Error processing arguments: {e}")
        sys.exit(1)


def main():
    if len(sys.argv) != 4:
        print("Usage: python get_args.py <tindex_values> <hours_back_to_test> <dt_values>")
        sys.exit(1)

    start_time = time.time()
    tindex_values_str = sys.argv[1]
    hours_back_to_test_str = sys.argv[2]
    dt_values_str = sys.argv[3]

    # Process arguments using the function - casting
    tindex_values, hours_back_to_test, dt_values = casting_arguments(tindex_values_str, hours_back_to_test_str, dt_values_str)

    batch_file_template = 'Batch_Control_History_template.xlsx'
    # tindex_values = [1.85,1.9] # manully off 
    jump_time = len(tindex_values)
    # hours_back_to_test = 6 #  # manully off
    size_of_value = len(tindex_values) * hours_back_to_test
    # dt_values = [0.0009,0.0008,0.0007,0.0006,0.0005,0.0004,0.0003,0.0002] # manully off
 
    big_dt_values = generate_big_tindex_values(dt_values, size_of_value)

    input_file = 'Batch_Control_History_1.xlsx'
    sheet_name = 'Sheet1'
    column_name = 'DT '

    number_of_lines = run_flow_build_batch_hitory(batch_file_template,jump_time,hours_back_to_test,dt_values)
    update_excel_column(input_file, sheet_name, column_name, big_dt_values)
    column_name = 'Tindex Duplicate  Point'
    update_excel_column(input_file, sheet_name, column_name, tindex_values)
    number_of_lines+=1 
    copy_file_to_same_folder(batch_file_template)

    script_dir = os.path.dirname(__file__)
    batch_control_history = os.path.join(script_dir, "Batch_Control_History_1.xlsx")
    batch_file_template = os.path.join(script_dir,"_Batch_Control_History_template.xlsx")

    process_files(batch_control_history, batch_file_template, sheet_name='Control Sheet', number_of_lines=number_of_lines)

    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f"Start time: {time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(start_time))}")
    print(f"End time: {time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(end_time))}")
    print(f"Function executed in {elapsed_time} seconds.")
        
    # Define the source file path, destination directory, and new file name
    src_file,dest_directory,new_file_name = copy_template_to_destination()
    rename_file_and_copy(src_file, dest_directory, new_file_name)

if __name__ == "__main__":
    main()