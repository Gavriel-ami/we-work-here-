import openpyxl
from datetime import datetime
from collections import Counter
from openpyxl import Workbook , load_workbook
import os
import sys
from openpyxl.utils import get_column_letter

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from helpers import excel_io
from send_email import send_email

def get_first_symbol(df):
    symbol_text = df['Sumbol'].loc[0]
    print(f"Building the calculation page ,based on symbol:{symbol_text}\n") 
    return symbol_text.replace("/","_")

def get_dts_by_succesfull_prediction(df):
    filtered_df_by_successfull_prediction = df[df['TruePriceOnPredx2DirectionColumn'] == 1]
    dts = filtered_df_by_successfull_prediction['DT '].tolist()
    print(f'All of the dts that was successfull {dts}\n\n') 
    return dts

def calculat_percentage_success(df):
    total_dts = df['DT '].value_counts()
    dts = get_dts_by_succesfull_prediction(df)
    print("Counts of each unique value in the 'DT ' column:")
    print(total_dts.index)
    dt_combinations_found = list(total_dts.index)
    total_per_dt = total_dts.to_list()
    result_dict = {key: value for key, value in zip(dt_combinations_found, total_per_dt)}
    print(result_dict,"\n")
    counter = Counter(dts)
    good_stactics = 0
    email_body = ""
    # Print the count of each unique value
    for value, count in counter.items():
        percentage = count/result_dict[value] * 100
        print(f"Value {value} avg:  {percentage}  %  appears {count} times")
        email_body += f"Value {value} avg: {percentage:.2f}% appears {count} times\n"
        if count/result_dict[value] * 100 > 70 :
            good_stactics += 1   

        # add here the counter if the avg is over 70 
        # print(f"Value {value} appears {count} times")
        # Do avg over the file all of 000.2 and the amount of 1 are on . 
    return counter, result_dict , good_stactics, email_body

def append_to_excel(counter, result_dict, filename):
    # Load or create the workbook
    try:
        workbook = load_workbook(filename)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active

    # Find the next available column
    next_col = sheet.max_column + 1

    # Write headers if the file is new
    if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(1, 1).value is None:
        sheet.cell(row=1, column=1, value="Value")
        sheet.cell(row=1, column=2, value="Average (%)")
        next_col = 4  # Set to 4 because we are adding initial columns

    # Append current date and time to the next column
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.cell(row=1, column=next_col, value=f"Run {next_col - 3} ({current_time})")

    counter_sorted = sorted(counter.items())

    for i, (value, count) in enumerate(counter_sorted, start=2):
        avg = (count / result_dict[value]) * 100
        sheet.cell(row=i, column=1, value=value)
        sheet.cell(row=i, column=next_col, value=f"{avg:.2f}%")

    # Freeze the first row
    sheet.freeze_panes = sheet['A2']

    # Adjust column widths to fit content
    for col in sheet.columns:
        max_length = 0
        column = get_column_letter(col[0].column)  # Get the column letter
        for cell in col:
            if cell.value is not None:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except Exception as e:
                    print(f"Error adjusting width for cell {cell.coordinate}: {e}")
        adjusted_width = max_length + 2
        sheet.column_dimensions[column].width = adjusted_width

    # Save the workbook
    workbook.save(filename)
    workbook.close()
    print(f"\nOutput appended to {filename}")

def build_output_folder():
    ## Specify the directory path
    directory = "percentage_results"

    # Check if the directory already exists
    if not os.path.exists(directory):
        # Create the directory
        os.makedirs(directory)
        print("Directory created successfully!")
    else:
        print("Directory already exists!")

    return directory + "/"


def main():
    batch_file  = sys.argv[1]  
    email_text  = sys.argv[2]  
    email_names = email_text.split(", ")
    # gavrielmn@gmail.com,pinchas@amiteam.net,yakov@amiteam.net,or@amiteam.net
    print(f"THIS IS THE {batch_file}")
    df = excel_io.read_the_batch_history_file(batch_file)
    symbol_text = get_first_symbol(df)
    counter, result_dict, good_stactics , email_body= calculat_percentage_success(df)
    print(f'counter:{counter}')
    print(f'result_dict:{result_dict}')
    
    # email_names = ["gavrielmn@gmail.com","pinchas@amiteam.net"]
    # send email for good stistics 
    if good_stactics >= 2:# 2
        print(good_stactics)
        subject = f"statistics are good {symbol_text}"
        body = email_body +f"\n Our latest data analysis shows a significant improvement in our statistics, indicating strong performance - {symbol_text}"
        for email_address in email_names:
            print('email was sent to:', email_address)
            email_sender = "gavriel@amiteam.net"
            email_password = "akmz mzqy okrc farh" # A token password ,Not you email password . 
            email_receiver = email_address
            # email_receiver = "yakov@amiteam.net"
            send_email(subject, body, email_sender, email_password, email_receiver)



    ###### append the data into file #################### 
    filename = f'output_{symbol_text}.xlsx'
    folder_name = build_output_folder()
    filename  = folder_name + filename
    append_to_excel(counter, result_dict, filename)    

if __name__ == "__main__":
    main()
