import openpyxl
from datetime import datetime
from collections import Counter
from openpyxl import Workbook, load_workbook
import os
import sys
from openpyxl.utils import get_column_letter
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from helpers import excel_io

def get_first_symbol(df):
    symbol_text = df['Sumbol'].loc[0]
    print(f"Building the calculation page based on symbol: {symbol_text}")
    return symbol_text.replace("/", "_")

def calculate_percentage_success(df, columns):
    results = {}
    for column in columns:
        total_dts = df['DT '].value_counts()
        dts = get_dts_by_successful_prediction(df, column)
        print(f"Counts of each unique value in the 'DT ' column for {column}: {total_dts.index}")
        dt_combinations_found = list(total_dts.index)
        total_per_dt = total_dts.to_list()
        result_dict = {key: value for key, value in zip(dt_combinations_found, total_per_dt)}
        print(f"Result dictionary for {column}: {result_dict}")
        counter = Counter(dts)
        for value, count in counter.items():
            print(f"Value {value} avg for {column}: {count / result_dict[value] * 100}% appears {count} times")
        results[column] = (counter, result_dict)
    return results

def get_dts_by_successful_prediction(df, column):
    filtered_df_by_successful_prediction = df[df[column] == 1]
    dts = filtered_df_by_successful_prediction['DT '].tolist()
    print(f'All of the dts that were successful for {column}: {dts}')
    return dts

def append_to_excel(results, filename):
    try:
        workbook = load_workbook(filename)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active

    # Write headers if the file is new
    if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(1, 1).value is None:
        sheet.cell(row=1, column=1, value="Value")
        sheet.cell(row=1, column=2, value="Average (%)")

    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Find the next available column for each set of results
    start_col = sheet.max_column + 1

    for column, (counter, result_dict) in results.items():
        sheet.cell(row=1, column=start_col, value=f"{column} ({current_time})")
        counter_sorted = sorted(counter.items())
        for i, (value, count) in enumerate(counter_sorted, start=2):
            avg = (count / result_dict[value]) * 100
            if sheet.cell(row=i, column=1).value is None:
                sheet.cell(row=i, column=1, value=value)
            sheet.cell(row=i, column=start_col, value=f"{avg:.2f}%")
        start_col += 1

    sheet.freeze_panes = sheet['A2']
    for col in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except Exception as e:
                    logging.error(f"Error adjusting width for cell {cell.coordinate}: {e}")
        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save(filename)
    workbook.close()
    print(f"Output appended to {filename}")

def build_output_folder():
    directory = "percentage_results"
    if not os.path.exists(directory):
        os.makedirs(directory)
        print("Directory created successfully!")
    else:
        print("Directory already exists!")
    return directory + "/"

def main():
    if len(sys.argv) < 3:
        print("Usage: script.py <batch_file>")
        sys.exit(1)

    # batch_file = sys.argv[1]
    # columns_arg = sys.argv[2]
    batch_file, columns_arg = sys.argv[1], sys.argv[2]
    columns = columns_arg.split(",")  # Convert the columns argument to a list
    print(f"Processing {batch_file} for columns {columns}")

    df = excel_io.read_the_batch_history_file(batch_file)
    symbol_text = get_first_symbol(df)
    results = calculate_percentage_success(df, columns)
    
    filename = f'output_{symbol_text}.xlsx'
    folder_name = build_output_folder()
    filename = folder_name + filename
    append_to_excel(results, filename)

if __name__ == "__main__":
    main()
