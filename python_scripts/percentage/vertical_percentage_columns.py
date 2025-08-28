import openpyxl
from datetime import datetime
from collections import Counter
from openpyxl import Workbook, load_workbook
import os
import sys
import logging
from openpyxl.utils import get_column_letter

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from helpers import excel_io

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Define the columns to process
COLUMNS = ['TruePriceOnPredxDirectionColumn', 'TruePriceOnPredx2DirectionColumn'] 

def get_first_symbol(df):
    symbol_text = df['Sumbol'].loc[0]
    logging.info(f"Building the calculation page based on symbol: {symbol_text}")
    return symbol_text.replace("/", "_")

def get_dts_by_successful_prediction(df, column):
    filtered_df_by_successful_prediction = df[df[column] == 1]
    dts = filtered_df_by_successful_prediction['DT '].tolist()
    logging.info(f'All of the dts that were successful for {column}: {dts}')
    return dts

def calculate_percentage_success(df, columns):
    results = {}
    for column in columns:
        total_dts = df['DT '].value_counts()
        dts = get_dts_by_successful_prediction(df, column)
        logging.info(f"Counts of each unique value in the 'DT ' column for {column}: {total_dts.index}")
        dt_combinations_found = list(total_dts.index)
        total_per_dt = total_dts.to_list()
        result_dict = {key: value for key, value in zip(dt_combinations_found, total_per_dt)}
        logging.info(f"Result dictionary for {column}: {result_dict}")
        counter = Counter(dts)
        for value, count in counter.items():
            logging.info(f"Value {value} avg for {column}: {count / result_dict[value] * 100}% appears {count} times")
        results[column] = (counter, result_dict)
    return results

def append_to_excel(results, filename):
    try:
        workbook = load_workbook(filename)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active

    # Determine where to start writing the new results
    start_row = sheet.max_row + 2

    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for column, (counter, result_dict) in results.items():
        # Write header for the column set
        sheet.cell(row=start_row, column=1, value=f"{column} ({current_time})")
        sheet.cell(row=start_row, column=2, value="Value")
        sheet.cell(row=start_row, column=3, value="Average (%)")

        counter_sorted = sorted(counter.items())
        for i, (value, count) in enumerate(counter_sorted, start=start_row + 1):
            avg = (count / result_dict[value]) * 100
            sheet.cell(row=i, column=2, value=value)
            sheet.cell(row=i, column=3, value=f"{avg:.2f}%")

        start_row = i + 2  # Leave a blank row after each set of results

    sheet.freeze_panes = sheet['A2']
    for col in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except Exception as e:
                    logging.error(f"Error adjusting width for cell {cell.coordinate}: {e}")
        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save(filename)
    workbook.close()
    logging.info(f"Output appended to {filename}")

def build_output_folder():
    directory = "percentage_results"
    if not os.path.exists(directory):
        os.makedirs(directory)
        logging.info("Directory created successfully!")
    else:
        logging.info("Directory already exists!")
    return directory + "/"

def main():
    if len(sys.argv) < 2:
        logging.error("Usage: script.py <batch_file>")
        sys.exit(1)

    batch_file = sys.argv[1]
    logging.info(f"Processing {batch_file} for columns {COLUMNS}")

    df = excel_io.read_the_batch_history_file(batch_file)
    symbol_text = get_first_symbol(df)
    results = calculate_percentage_success(df, COLUMNS)
    filename = f'output_{symbol_text}.xlsx'
    folder_name = build_output_folder()
    filename = folder_name + filename
    append_to_excel(results, filename)

if __name__ == "__main__":
    main()
